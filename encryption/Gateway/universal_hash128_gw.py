import binascii
import csv
import hashlib
import math
import subprocess
import time
import numpy as np
import openpyxl
import argparse
import os
from hashlib import sha512
from tempfile import TemporaryFile
import xlrd
from xlwt import Workbook
import pyaes

def hash_function(input_data, hash_size):
    hash_object = sha512(input_data.encode())
    return [int(bit) for bit in bin(int(hash_object.hexdigest(), 16))[2:].zfill(hash_size)]

# Parse command-line arguments
parser = argparse.ArgumentParser(description='Data encryption check using SHA-128.')
parser.add_argument('--datapath', required=True, help='Path to the reconsiliation directory')
parser.add_argument('--destination', required=True, help='Path for the result destination')
args = parser.parse_args()

start4 = time.time()

# Read data from file
ws = []
with open(os.path.join(args.datapath,'Decoding_Gateway_Tanpa_Parity.csv')) as f:
    for row in csv.reader(f):
        ws.append(row)
data = [elements for ubah in ws for elements in ubah]

# Create a concatenated string of data and convert to a list of integers
alice1 = ''.join(data)
bitalice = list(map(int, alice1))

# Save to Excel for consistency
book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = 'HasilBCH'
sheet1.cell(row=1, column=1, value='Apgateway')
for i in range(1, len(bitalice) + 1):
    sheet1.cell(row=i + 1, column=1, value=int(bitalice[i - 1]))
book.save('decodinggateway_tanpaparity.xlsx')

# Read Excel and extract bit values
workbook = openpyxl.load_workbook('decodinggateway_tanpaparity.xlsx', read_only=True)
worksheet = workbook.active
bbob = [int(worksheet.cell(row=row, column=1).value) for row in range(2, worksheet.max_row + 1)]

# Determine the number of keys to be generated
jumlahkeya = math.floor(len(bbob) / 128)
jumlahkey = min(48, jumlahkeya)
ukuranhash = 128
aaaa = len(bbob) % 128
lenbob = len(bbob) - aaaa

for i in range(aaaa):
    del bbob[lenbob]

Hashtab = []
with open('Hashtable128.csv', newline='') as f:
    reader = csv.reader(f)
    for row in reader:
        Hashtab.append(row)
Hashtable = np.array(Hashtab)

key2 = []
for i in range(jumlahkey):
    mat2 = []
    bbb = bbob[i * ukuranhash:(ukuranhash * (i + 1))]
    for x in range(len(Hashtable)):
        row = []
        total = 0
        for y in range(len(bbb)):
            total = total + (int(Hashtable[x][y]) * bbb[y])
            row = total % 2
        mat2.append(int(row))
    key2.append(mat2)

v = 0
bx = [0] * 128 * jumlahkey
for i in range(jumlahkey):
    for j in range(128):
        bx[v] = key2[i][j]
        v = v + 1

# Save the key to Excel
book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = 'univhash'
sheet1.cell(row=1, column=1, value='gateway')
for i in range(1, len(bx) + 1):
    sheet1.cell(row=i + 1, column=1, value=int(bx[i - 1]))
book.save('gatewayUnivHash128.xlsx')

# Convert the key to CSV format
univ = [bx[i] for i in range(len(bx))]
with open('hashok128.csv', 'w', newline='') as fp:
    a = csv.writer(fp, delimiter=',')
    a.writerows([[bit] for bit in univ])

end4 = time.time()
waktu_hash = end4 - start4
print('-------------------')
print('Universal Hash Berhasil')
print('-------------------')

startnist = time.time()
# command = 'gcc -o "C:/Users/MHK/Documents/Folder Izzat/rssi-generator/encryption/NIST-Test128" "C:/Users/MHK/Documents/Folder Izzat/rssi-generator/encryption/NIST-Test128.c"'
# subprocess.run(command, shell=True)

command = "E:/My Code Projects/rssi_generator/encryption/Node/NIST-Test128.exe"
subprocess.run(command, shell=True)

indeks = []
indek = []

time.sleep(1)

with open("NISTHash1.csv", newline="") as f:
    reader = csv.reader(f)
    for row in reader:
        indeks.append(row)
prindek = []
for i in range(0, len(indeks)):
    indek.append(int(indeks[i][0]))
    prindek.append(int(indeks[i][0]) + 1)
endnist = time.time()
print('===========~~~~~~~~~~~=========~~~~~~~~~~~~~~==============\n')
print('NIST Hasil prioritas index',prindek)
print('Waktu Proses Uji NIST : ', endnist - startnist)

start6 = time.time()

databob=[]
hex2=[]
for j in range(len(indek)):
    hash2=[0]*128
    for i in range(128):
        hash2[i]=key2[indek[j]][i]

    data2=''.join(str(e) for e in hash2)
    someText2 = data2.encode("ascii")
    b2=hashlib.sha1(someText2).hexdigest()
    hex2.append(b2)

    print ('Hasil hash AP bob Kunci-{} = {}'.format(indek[j]+1,hex2[j]))

#time.sleep(3)

book=Workbook()
sheet1=book.add_sheet('sha')
sheet1.write(0,0,'gateway')
for i in range (1,len(hex2)+1):
    sheet1.write(i,0,hex2[i-1])
book.save('Hashgateway.xls')
book.save(TemporaryFile())


workbook = xlrd.open_workbook('E:/My Code Projects/rssi_generator/encryption/Node/Hashnode.xls', on_demand = True)
worksheet = workbook.sheet_by_index(0)
first_row = [] # Header
for col in range(0,worksheet.ncols):
    first_row.append( worksheet.cell_value(0,col) )
    # tronsform the workbook to a list of dictionnaries
hex1 = []
for row in range(1, worksheet.nrows):
    elm = {}
    for col in range(1):
        elm=worksheet.cell_value(row,col)
    hex1.append(elm)

for j in range(len(indek)):
    if(hex1[j]==hex2[j]):
        print('Hash Value-%d valid, proses enkripsi dapat dilakukan'%(j+1))
        (j+1)
    else:
        print('Hash Value-%d not valid'%(j+1))

end6 = time.time()
waktu_SHA = end6-start6
print('-------------------')
print('SHA Berhasil')
print('-------------------')

print('\n\n===========~~~~~~~~~~~== AES ~~~~~~~~~~~~~~==============')
print('===========~~~~~~~~~~~=========~~~~~~~~~~~~~~==============\n')    
start_aes=time.time()
for kuncinya in range(len(indek)):
    if(hex1[indek[kuncinya]]==hex2[indek[kuncinya]]):
        # =========================================================================
        # ================================= AES-256 ===============================

        keybit = ''.join(str(e) for e in key2[indek[kuncinya]])
        keyint=int(keybit,2)
        keybyte=binascii.unhexlify('%x' % keyint)
##        keyhex=keybyte.hex()
        print('Key Bob (int)',keybit)
        print('Key Bob (byte)',keybyte)
##        print('Key Bob (hex)',keyhex)
        with open("keyA.txt", "wb") as binary_file:
            writen = binary_file.write(keybyte)
    
        #=======================================================CTR===
        print ('\n=================CTR=============')
        ciphertext= b'\xa8r\xd1\x7f\xb48\xa1\x17\xe3QJ\x0e"0\x99w'

        # show the encrypted data
        print ('ciphertext = ',ciphertext)

        # DECRYPTION
        # CTR mode decryption requires a new instance be created
        aesctr = pyaes.AESModeOfOperationCTR(keybyte)

        # decrypted data is always binary, need to decode to plaintext
        decrypted = aesctr.decrypt(ciphertext)
        
        # True
        print ('Dekripsi BERHASIL\nDecrypted cipher = ',decrypted)
        time.sleep(4)
        print('Proses selesai')
        break
    elif(hex1[kuncinya]!=hex2[kuncinya] and (kuncinya+1)<(len(indek)+1)):
        print('Hash untuk kunci ke-%d tidak valid, maka pakai kunci selanjutnya (Kunci ke-%d)'%(indek[kuncinya]+1,indek[kuncinya+1]+1))
    else:
        print('Hash untuk kunci ke-%d tidak valid, ulangi proses dari awal (Pengukuran)'%(indek[kuncinya]+1))
        break
end_aes=time.time()
waktu_aes=end_aes-start_aes
