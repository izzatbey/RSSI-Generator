import csv
import math
import time
import numpy as np
import openpyxl
import argparse
import os
from hashlib import sha512

def hash_function(input_data, hash_size):
    hash_object = sha512(input_data.encode())
    return [int(bit) for bit in bin(int(hash_object.hexdigest(), 16))[2:].zfill(hash_size)]

# Parse command-line arguments
parser = argparse.ArgumentParser(description='Data encryption check using SHA-128.')
parser.add_argument('--datapath', required=True, help='Path to the reconsiliation directory')
parser.add_argument('--destination', required=True, help='Path for the result destination')
args = parser.parse_args()

start4 = time.time()

# Read data from file
ws = []
with open(os.path.join(args.datapath,'Decoding_Node_Tanpa_Parity.csv')) as f:
    for row in csv.reader(f):
        ws.append(row)
data = [elements for ubah in ws for elements in ubah]

# Create a concatenated string of data and convert to a list of integers
alice1 = ''.join(data)
bitalice = list(map(int, alice1))

# Save to Excel for consistency
book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = 'HasilBCH'
sheet1.cell(row=1, column=1, value='ApNode')
for i in range(1, len(bitalice) + 1):
    sheet1.cell(row=i + 1, column=1, value=int(bitalice[i - 1]))
book.save(os.path.join(args.destination,'decodingnode_tanpaparity.xlsx'))

# Read Excel and extract bit values
workbook = openpyxl.load_workbook(os.path.join(args.destination,'decodingnode_tanpaparity.xlsx'), read_only=True)
worksheet = workbook.active
bbob = [int(worksheet.cell(row=row, column=1).value) for row in range(2, worksheet.max_row + 1)]

# Determine the number of keys to be generated
jumlahkeya = math.floor(len(bbob) / 128)
jumlahkey = min(48, jumlahkeya)
ukuranhash = 128
aaaa = len(bbob) % 128
lenbob = len(bbob) - aaaa

for i in range(aaaa):
    del bbob[lenbob]

Hashtab = []
with open(os.path.join(args.destination,'Hashtable128.csv'), newline='') as f:
    reader = csv.reader(f)
    for row in reader:
        Hashtab.append(row)
Hashtable = np.array(Hashtab)

key1 = []
for i in range(jumlahkey):
    mat2 = []
    bbb = bbob[i * ukuranhash:(ukuranhash * (i + 1))]
    for x in range(len(Hashtable)):
        row = []
        total = 0
        for y in range(len(bbb)):
            total = total + (int(Hashtable[x][y]) * bbb[y])
            row = total % 2
        mat2.append(int(row))
    key1.append(mat2)

v = 0
bx = [0] * 128 * jumlahkey
for i in range(jumlahkey):
    for j in range(128):
        bx[v] = key1[i][j]
        v = v + 1

# Save the key to Excel
book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = 'univhash'
sheet1.cell(row=1, column=1, value='Node')
for i in range(1, len(bx) + 1):
    sheet1.cell(row=i + 1, column=1, value=int(bx[i - 1]))
book.save(os.path.join(args.destination,'NodeUnivHash128.xlsx'))

# Convert the key to CSV format
univ = [bx[i] for i in range(len(bx))]
with open(os.path.join(args.destination,'hashok128.csv'), 'w', newline='') as fp:
    a = csv.writer(fp, delimiter=',')
    a.writerows([[bit] for bit in univ])

end4 = time.time()
waktu_hash = end4 - start4
print('-------------------')
print('Universal Hash Berhasil')
print('-------------------')
