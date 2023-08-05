import math
import time
from openpyxl import Workbook
import xlrd
import numpy as np
from math import log
from tempfile import TemporaryFile

from comm.socket_comm import socketrecv, socketsend, init_socket

print("\n============================ Jana Multibit =======================")
mulai=time.time()
start_jana=time.time()
workbook = xlrd.open_workbook('E:/rssi_generator/datasets/Bob_Praproses.xls', on_demand = True)
worksheet = workbook.sheet_by_index(0)
##print worksheet.ncols		#perintah untuk mengetahui jumlah column
nilaiasli = []
for row in range(1, worksheet.nrows):
	value = worksheet.cell_value(row,0)
	nilaiasli.append(value)
nilaiasli=np.array(nilaiasli)

jana_alice=[]
jana_alice1=[]
for i in range(len(nilaiasli)):
	jana_alice.append(str(nilaiasli[i]))
nilailebih = (len(jana_alice)%128)
jmlblok= math.floor(len(jana_alice)/128)
for i in range(0,(len(jana_alice)-nilailebih)):
	jana_alice1.append(str(jana_alice[i]))
	
# jana_alice=[] #mengosongkan variabel

# for blok in range(0,jmlblok):
	# tmp=[]
	# if blok == 0:
		# tmp.extend(jana_alice1[0:128])
	# else:
		# tmp.extend(jana_alice1[128*blok:128*(blok+1)])
	# jana_alice.append(tmp)
	
jana_alice=[] #mengosongkan variabel

for blok in range(0,jmlblok):
	tmp=[]
	if blok == 0:
		tmp.extend(jana_alice1[0:128])
	else:
		tmp.extend(jana_alice1[128*blok:128*(blok+1)])
	jana_alice.append(tmp)
del jana_alice1
hasiljana=[]
for coba in range(0,len(jana_alice)):
	jana_alice1=[]
	jana_alice1.extend(jana_alice[coba])
	L = [ (jana_alice1[i],i) for i in range(len(jana_alice1)) ]	##addressing array
	L.sort()
	nilairss,addressing = zip(*L)
	#print ("\n\nPanjang Data awal \n",(len(nilairss)))		#melihat panjang data
	p = (len(nilairss))
	N = log(p,2)
	a = 2
	M = 2**a
	if a <= N:
		M = 2**a

	bb = M						#menginginkan 4 kolom
	filter=(len(nilairss)%bb)							#Memfilter jumlah data
	if (filter != 0):									#agar dapat di proses
		pengurangan = (len(nilairss)-filter)			#dan tetap menggunakan
		nilairss1=[]									#jumlah kolom yang diinginkan
		for i in range(0,pengurangan):
			nilairss1.append(nilairss[i])
		nilairss1 = np.array(nilairss1)
	else :
		nilairss1 = np.array(nilairss)
	aa = int(len(nilairss1)/bb)
	x=nilairss1.reshape(bb,aa).T
	#print(x)
	#print ("\n\nDirubah menjadi\n\n")
	##==pemberian angka==##
	for i in range(0,bb):
		for j in range (0,(len(x))):
			if i == 0:
				x[j][i] = 4
			elif i == 1:
				x[j][i] = 3
			elif i == 2:
				x[j][i] = 2
			elif i == 3:
				x[j][i] = 1
	#print (x)
	x1=x.T
	x2=x1.reshape(aa*bb,1)					##dikembalikan ke array 1 kolom
	bentukawal=[]
	for i in range(len(x2)):
		bentukawal.append(int(x2[i]))
	unsort = list(zip(addressing,bentukawal))		#proses pengembalian
	unsort.sort()							#sesuai address
	addressing,bitbit = zip(*unsort)
	hasil = []
	for i in range(len(bitbit)):
		if bitbit[i]==4:
			hasil.append(1)
			hasil.append(0)
		elif bitbit[i]==3:
			hasil.append(1)
			hasil.append(1)
		elif bitbit[i]==2:
			hasil.append(0)
			hasil.append(1)
		elif bitbit[i]==1:
			hasil.append(0)
			hasil.append(0)
	hasiljana.append(hasil)
	print ("hasil kuantisasi \n",hasil)
	print ("jumlah key sekarang ",len(hasil))
# janatime=time.time()
#convert 1 array
hasilakhir=[]
for i in range(0,len(hasiljana)):
	for j in range(0,len(hasiljana[i])):
		hasilakhir.append(hasiljana[i][j])
book=Workbook()
sheet1 = book.create_sheet('kuantisasiALICE')
sheet1.cell(row=1, column=1, value='AliceKuan')
for i in range(1,len(hasilakhir)+1):
	sheet1.cell(row=i+1, column=1, value=int(hasilakhir[i-1]))
book.save('E:/rssi_generator/datasets/hasilkuantisasiBob.xls')
book.save(TemporaryFile())
end_jana=time.time()
time_jana=end_jana-start_jana
kgrkuan=len(hasilakhir)*(end_jana-start_jana)
print('KGR = %f'%kgrkuan)
print('waktu komputasi kuantisasi = {}seconds'.format(end_jana-start_jana))
socketrecv(init_socket())
socketsend('hasilkuantisasiBob.xls', init_socket())
print("Kuantisasi BOB Berhasil")