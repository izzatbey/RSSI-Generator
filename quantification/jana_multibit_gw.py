import math
import argparse
import os
import time
from openpyxl import Workbook
import xlrd
import numpy as np
from math import log
from tempfile import TemporaryFile

from comm.socket_comm import socketrecv, socketsend, init_socket

def load_data(file_path):
    workbook = xlrd.open_workbook(file_path, on_demand=True)
    worksheet = workbook.sheet_by_index(0)
    data = []
    for row in range(1, worksheet.nrows):
        value = worksheet.cell_value(row, 0)
        data.append(str(value))
    return np.array(data)

def process_data(data):
    jana_data = []
    jana_data1 = []
    
    for value in data:
        jana_data.append(value)
    
    block_length = 128
    excess_length = len(jana_data) % block_length
    num_blocks = math.floor(len(jana_data) / block_length)
    
    for i in range(0, len(jana_data) - excess_length):
        jana_data1.append(jana_data[i])
    
    jana_data = []
    
    for block in range(0, num_blocks):
        tmp = []
        if block == 0:
            tmp.extend(jana_data1[0:block_length])
        else:
            tmp.extend(jana_data1[block_length * block:block_length * (block + 1)])
        jana_data.append(tmp)
    
    hasil_jana = []
    for coba in range(0, len(jana_data)):
        jana_data1 = []
        jana_data1.extend(jana_data[coba])
        L = [(jana_data1[i], i) for i in range(len(jana_data1))]
        L.sort()
        nilairss, addressing = zip(*L)
        
        p = len(nilairss)
        N = log(p, 2)
        a = 2
        M = 2 ** a
        if a <= N:
            M = 2 ** a
        
        bb = M
        filter_len = len(nilairss) % bb
        if filter_len != 0:
            reduction = len(nilairss) - filter_len
            nilairss1 = np.array(nilairss[:reduction])
        else:
            nilairss1 = np.array(nilairss)
        
        aa = int(len(nilairss1) / bb)
        x = nilairss1.reshape(bb, aa).T
        
        for i in range(0, bb):
            for j in range(0, len(x)):
                if i == 0:
                    x[j][i] = 4
                elif i == 1:
                    x[j][i] = 3
                elif i == 2:
                    x[j][i] = 2
                elif i == 3:
                    x[j][i] = 1
        
        x1 = x.T
        x2 = x1.reshape(aa * bb, 1)
        bentukawal = [int(val) for val in x2]
        unsorted = list(zip(addressing, bentukawal))
        unsorted.sort()
        addressing, bitbit = zip(*unsorted)
        
        hasil = []
        for bit in bitbit:
            if bit == 4:
                hasil.extend([1, 0])
            elif bit == 3:
                hasil.extend([1, 1])
            elif bit == 2:
                hasil.extend([0, 1])
            elif bit == 1:
                hasil.extend([0, 0])
        hasil_jana.append(hasil)
    
    return hasil_jana

def save_results(results, file_path):
    book = Workbook()
    sheet1 = book.create_sheet('kuantisasiBOB')
    sheet1.cell(row=1, column=1, value='BobKuan')
    for i, result in enumerate(results):
        sheet1.cell(row=i + 2, column=1, value=int(result))
    book.save(file_path)
    book.save(TemporaryFile())

mulai = time.time()
start_jana = time.time()

# Parse command-line arguments
parser = argparse.ArgumentParser(description='Data quantification using Jana Multibit process.')
parser.add_argument('--datapath', required=True, help='Path to the preprocess directory')
parser.add_argument('--filename', required=True, help='filename of the preprocess')
parser.add_argument('--destination', required=True, help='Path for the result destination')
args = parser.parse_args()

data_file_path = os.path.join(args.datapath, args.filename)


data = load_data(data_file_path)
hasil_jana = process_data(data)

hasil_akhir = []
for hasil_jana_item in hasil_jana:
    hasil_akhir.extend(hasil_jana_item)

hasil_file_path = os.path.join(args.destination, 'hasilkuantisasiBob.xls') 
save_results(hasil_akhir, hasil_file_path)

end_jana = time.time()
time_jana = end_jana - start_jana
kgr_kuan = len(hasil_akhir) * (end_jana - start_jana)

print('KGR = %f' % kgr_kuan)
print('waktu komputasi kuantisasi = {} seconds'.format(end_jana - start_jana))

socketrecv(init_socket())
socketsend(hasil_file_path, init_socket())

print("Kuantisasi BOB Berhasil")
