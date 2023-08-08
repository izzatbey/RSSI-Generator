import math
import argparse
import os
import time
import numpy as np
from openpyxl import Workbook
import xlrd
from math import log
from tempfile import TemporaryFile

from comm.socket_comm import socketrecv, socketsend, init_socket

def load_data_from_excel(file_path):
    workbook = xlrd.open_workbook(file_path, on_demand=True)
    worksheet = workbook.sheet_by_index(0)
    nilaiasli = [worksheet.cell_value(row, 0) for row in range(1, worksheet.nrows)]
    return np.array(nilaiasli)

def quantize_data(data, block_size=128):
    jana_alice = [str(value) for value in data]
    nilailebih = len(jana_alice) % block_size
    jmlblok = math.floor(len(jana_alice) / block_size)
    jana_alice1 = [str(jana_alice[i]) for i in range(len(jana_alice) - nilailebih)]
    
    quantized_data = []
    for blok in range(0, jmlblok):
        tmp = jana_alice1[128 * blok:128 * (blok + 1)] if blok != 0 else jana_alice1[0:128]
        quantized_data.append(tmp)
    
    return quantized_data

def perform_quantization(quantized_data):
    hasiljana = []
    for jana_alice1 in quantized_data:
        L = [(jana_alice1[i], i) for i in range(len(jana_alice1))]
        L.sort()
        nilairss, addressing = zip(*L)
        p = len(nilairss)
        N = log(p, 2)
        a = 2
        M = 2 ** a if a <= N else 2
        
        bb = M
        filter = len(nilairss) % bb
        if filter != 0:
            pengurangan = len(nilairss) - filter
            nilairss1 = np.array(nilairss[:pengurangan])
        else:
            nilairss1 = np.array(nilairss)
            
        aa = int(len(nilairss1) / bb)
        x = nilairss1.reshape(bb, aa).T
        
        for i in range(bb):
            for j in range(len(x)):
                x[j][i] = 4 - i
        
        x1 = x.T
        x2 = x1.reshape(aa * bb, 1)
        bentukawal = [int(x2[i]) for i in range(len(x2))]
        
        unsort = sorted(zip(addressing, bentukawal))
        addressing, bitbit = zip(*unsort)
        
        hasil = []
        for i in range(len(bitbit)):
            if bitbit[i] == 4:
                hasil.extend([1, 0])
            elif bitbit[i] == 3:
                hasil.extend([1, 1])
            elif bitbit[i] == 2:
                hasil.extend([0, 1])
            elif bitbit[i] == 1:
                hasil.extend([0, 0])
        
        hasiljana.append(hasil)
        print("hasil kuantisasi \n", hasil)
        print("jumlah key sekarang ", len(hasil))
    
    return hasiljana

def save_quantized_data(quantized_data, file_path):
    book = Workbook()
    sheet1 = book.create_sheet('kuantisasiALICE')
    sheet1.cell(row=1, column=1, value='AliceKuan')
    for i, data in enumerate(quantized_data):
        for j in range(len(data)):
            sheet1.cell(row=i * len(data) + j + 2, column=1, value=int(data[j]))
    book.save(file_path)
    book.save(TemporaryFile())

mulai = time.time()
start_jana = time.time()

# Parse command-line arguments
parser = argparse.ArgumentParser(description='Data quantification using Jana Multibit process.')
parser.add_argument('--datapath', required=True, help='Path to the preprocess directory')
parser.add_argument('--filename', required=True, help='filename of the preprocess')
parser.add_argument('--destination', required=True, help='Path for the result destination')
args = parser.parse_args()

data_file_path = os.path.join(args.datapath, args.filename)

preprocess_data = load_data_from_excel(data_file_path)

quantized_data = quantize_data(preprocess_data)
hasiljana = perform_quantization(quantized_data)

hasilakhir = [bit for data in hasiljana for bit in data]

hasil_file_path = os.path.join(args.destination, 'hasilkuantisasiAlice.xls') 
save_quantized_data(hasiljana, hasil_file_path)

end_jana = time.time()
time_jana = end_jana - start_jana
kgrkuan = len(hasilakhir) * (end_jana - start_jana)

print('KGR = %f' % kgrkuan)
print('waktu komputasi kuantisasi = {} seconds'.format(end_jana - start_jana))

socketsend(hasil_file_path, init_socket())
socketrecv(init_socket())
print("Kuantisasi ALICE Berhasil")
